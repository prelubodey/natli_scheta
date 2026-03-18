import asyncio
import json
import logging
import os
from datetime import datetime, timedelta

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command

import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Читаем токены СТРОГО из переменных окружения (файла .env)
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if not TELEGRAM_BOT_TOKEN or not GEMINI_API_KEY:
    raise ValueError("Отсутствуют токены в переменных окружения. Пожалуйста, создайте файл .env или запустите скрипт установки (setup.bat / setup.sh).")

genai.configure(api_key=GEMINI_API_KEY)

# Используем модель gemini-2.5-flash
model = genai.GenerativeModel('gemini-2.5-flash', generation_config={"response_mime_type": "application/json"})

bot = Bot(token=TELEGRAM_BOT_TOKEN)
dp = Dispatcher()
logging.basicConfig(level=logging.INFO)

# Словарь для хранения данных пользователей (в реальном проекте лучше использовать БД)
# Формат: { user_id: [ { "Контрагент": "...", "№ счет": "...", "Организация": "...", "Подразделение": "...", "Дата": "DD.MM.YYYY", "Сумма": "..." } ] }
user_data_store = {}

PROMPT = """
Извлеки из этого документа (счета) данные и верни их строго в формате JSON с ключами:
- "supplier": Наименование поставщика (контрагент). Обязательно сокращай организационно-правовую форму до аббревиатур: вместо "ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ" пиши "ООО", вместо "Индивидуальный предприниматель" - "ИП", "АО", "ПАО" и т.д. Без кавычек в самом названии, если можно (например ООО Новая жизнь).
- "invoice_number": Номер счета (только цифры и буквы, например 14295 или №14295).
- "buyer": Наименование покупателя (организация). Также обязательно сокращай форму до ООО, ИП и т.д.
- "department": Подразделение (найди рукописную надпись на счете, например "цех"). Если нет - пустая строка.
- "date": Дата счета в формате DD.MM.YYYY (например 10.03.2026).
- "amount": Итоговая сумма по счету (только число с пробелами, например "10 060,00" или "10060").

Ответь ТОЛЬКО JSON-объектом, без пояснений и markdown-разметки.
"""

def process_and_add_data(user_id, extracted_data):
    if user_id not in user_data_store:
        user_data_store[user_id] = []
        
    supplier = extracted_data.get("supplier", "Не найдено").upper()
    invoice_number = extracted_data.get("invoice_number", "Не найдено")
    buyer = extracted_data.get("buyer", "Не найдено").upper()
    department = extracted_data.get("department", "Не найдено").upper()
    date_str = extracted_data.get("date", "01.01.2026")
    amount = extracted_data.get("amount", "0")
    
    # Пытаемся распарсить дату
    try:
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
        pay_date_obj = date_obj + timedelta(days=14)
        pay_date_str = pay_date_obj.strftime("%d.%m.%Y")
    except ValueError:
        pay_date_str = "Ошибка даты"
        
    record = {
        "Контрагент": supplier,
        "№ счет": invoice_number,
        "Организация": buyer,
        "Подразделение": department,
        "Дата": date_str,
        "Сумма": amount,
        "Срок отсрочк": "14",
        "Срок оплаты": pay_date_str,
        "Дата оплат": "" # всегда пусто
    }
    
    user_data_store[user_id].append(record)
    return record


@dp.message(Command("start"))
async def cmd_start(message: Message):
    await message.answer(
        "Привет! Отправь мне фото или PDF счета на оплату.\n"
        "Я распознаю данные, извлеку нужные поля (включая рукописное подразделение) и сохраню.\n\n"
        "Когда закончишь отправлять счета, напиши команду /tab, и я сформирую красивую таблицу Excel (Платежный календарь)."
    )

async def handle_file(message: Message, file_id: str, file_name: str):
    user_id = message.from_user.id
    processing_msg = await message.answer("Файл получен, отправляю на распознавание ИИ...")
    
    file_path = f"temp_{file_id}_{file_name}"
    
    try:
        # Скачиваем файл из телеграма
        file = await bot.get_file(file_id)
        await bot.download_file(file.file_path, file_path)
        
        # Загружаем файл в Gemini
        uploaded_file = genai.upload_file(file_path)
        
        # Генерируем ответ
        response = model.generate_content([uploaded_file, PROMPT])
        
        # Удаляем файл из Gemini
        uploaded_file.delete()
        
        try:
            extracted_json = json.loads(response.text)
        except json.JSONDecodeError:
            await processing_msg.edit_text(f"Ошибка парсинга JSON от ИИ. Ответ: {response.text}")
            return
            
        record = process_and_add_data(user_id, extracted_json)
        
        answer_text = (
            f"Документ распознан\n"
            f"Контрагент - {record['Контрагент']}\n"
            f"Счет №{record['№ счет']}\n"
            f"Организация - {record['Организация']}\n"
            f"Подразделение - {record['Подразделение']}\n"
            f"Дата - {record['Дата']}\n"
            f"Сумма - {record['Сумма']}\n"
            f"Отсрочка - {record['Срок отсрочк']}\n"
            f"Срок оплаты - {record['Срок оплаты']}\n"
            f"Дата оплаты -\n\n"
            f"Отправь следующий счет или нажми /tab для получения таблицы."
        )
        await processing_msg.edit_text(answer_text)
        
    except Exception as e:
        await processing_msg.edit_text(f"Произошла ошибка при обработке: {e}")
    finally:
        # Удаляем временный файл с диска
        if os.path.exists(file_path):
            os.remove(file_path)

@dp.message(F.photo)
async def handle_photo(message: Message):
    photo = message.photo[-1] # берем самое большое разрешение
    await handle_file(message, photo.file_id, "photo.jpg")

@dp.message(F.document)
async def handle_document(message: Message):
    doc = message.document
    if doc.mime_type == 'application/pdf' or doc.file_name.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
        await handle_file(message, doc.file_id, doc.file_name)
    else:
        await message.answer("Пожалуйста, отправляйте только PDF документы или изображения.")

@dp.message(Command("tab"))
async def cmd_tab(message: Message):
    user_id = message.from_user.id
    records = user_data_store.get(user_id, [])
    
    if not records:
        await message.answer("Вы еще не загрузили ни одного счета. Сначала отправьте фото или PDF.")
        return
        
    await message.answer("Формирую таблицу...")
    
    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Платежный календарь"

    # Заголовок
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "Платежный календарь"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Шапка
    headers = ["№", "Контрагент", "№ счет", "Организация", "Подразделение", "Дата", "Сумма", "Срок отсрочк", "Срок оплаты", "Дата оплат"]
    ws.append(headers)

    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = thin_border

    # Вспомогательная функция парсинга чисел и дат для Excel
    def parse_date_for_excel(d_str):
        if not d_str or d_str.strip() == "":
            return None
        try:
            return datetime.strptime(d_str, "%d.%m.%Y").date()
        except ValueError:
            return d_str

    def parse_float_for_excel(amount_str):
        # Удаляем пробелы и заменяем запятую на точку для Python
        cln = amount_str.replace(" ", "").replace(",", ".")
        try:
            return float(cln)
        except ValueError:
            return amount_str

    # Заполняем данные
    for i, record in enumerate(records):
        row_idx = 3 + i
        
        # Разбираем значения в нужные типы
        date_val = parse_date_for_excel(record["Дата"])
        pay_date_formula = f"=F{row_idx}+H{row_idx}"
        paid_date_val = parse_date_for_excel(record["Дата оплат"])
        
        amount_val = parse_float_for_excel(record["Сумма"])
        delay_val = 14
        
        row_data = [
            "", # Колонка номер по порядку всегда пустая (A)
            record["Контрагент"], # B
            record["№ счет"], # C
            record["Организация"], # D
            record["Подразделение"], # E
            date_val, # F
            amount_val, # G
            delay_val, # H
            pay_date_formula, # I
            paid_date_val # J
        ]
        
        ws.append(row_data)
        
        # Настраиваем форматирование ячеек добавленной строки
        for col_idx in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Формат даты
            if col_idx in [6, 9, 10]:
                cell.number_format = 'DD.MM.YYYY'
            # Формат суммы (число с 2 знаками и разделителем разрядов)
            elif col_idx == 7:
                cell.number_format = '#,##0.00'
            # Формат отсрочки
            elif col_idx == 8:
                cell.number_format = '0'

    # Ширина колонок
    column_widths = {
        'A': 6, 'B': 40, 'C': 15, 'D': 15, 'E': 25, 
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    file_name = f"Платежный_календарь_{user_id}.xlsx"
    wb.save(file_name)
    
    # Отправляем файл пользователю
    excel_file = FSInputFile(file_name)
    await bot.send_document(message.chat.id, excel_file, caption="Ваша таблица готова!")
    
    # Очищаем данные пользователя и удаляем файл
    user_data_store[user_id] = []
    os.remove(file_name)

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
